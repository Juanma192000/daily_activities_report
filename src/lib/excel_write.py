import pandas as pd 
from lib.activities_api import get_mettings
from datetime import date
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from enum import Enum
from config.config import EXCEL_PATH_READ, NAME_COLABORADOR, TOTAL_HOURS

class Element(Enum):
    ALIGMENT = 1
    BORDER = 2
    WB = 3
    
items=[
    "Title",
    "Proyecto",
    "Fecha",
    "Total de Hrs",
    "Ticket Jira"
    ]
def create_file():        
    writer = pd.ExcelWriter(EXCEL_PATH_READ)
    workbook  = writer.book    
    merge_format = workbook.add_format({
        'bold': 2,
        'border': 2,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#eeeeee'})
    merge_format.set_font_size(20)
    header_bold_format = workbook.add_format({
                'bold': 2,
                'border': 2,
                'align': 'center',
                'valign': 'vcenter'})
    content_format = workbook.add_format({
                'bold': False,
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'})
        ###### New Worksheet ######
    worksheet = workbook.add_worksheet(f"Cargos_{NAME_COLABORADOR}")
    worksheet.merge_range(0, 0, 1, 4, "Trabajo remoto", merge_format)

    return writer,worksheet,header_bold_format,content_format

def get_element(element):
    if element == Element.WB:
        return load_workbook(EXCEL_PATH_READ)
    elif element == Element.ALIGMENT:
        return Alignment(horizontal='center', vertical='center')
    elif element == Element.BORDER:
        return Border(left=Side(border_style='thin', color='000000'),
               right=Side(border_style='thin', color='000000'),
               top=Side(border_style='thin', color='000000'),
               bottom=Side(border_style='thin', color='000000'))
    else:
        return None
    
def fill_excel_with_meetings():
    try:
        print("Opening File...")    
        pd.read_excel(EXCEL_PATH_READ,index_col=None,engine='openpyxl')  
        wb = get_element(Element.WB)
        alignment = get_element(Element.ALIGMENT)
        borde = get_element(Element.BORDER)
        ws = wb.active
        ws.hidden = False
        mettings=get_mettings()  
        total_rows = ws.max_row
        index=total_rows+1
        for meeting in mettings:
            try:
                ws.cell(row=index, column=1).value = meeting.title
                ws.cell(row=index, column=2).value = meeting.proyect
                ws.cell(row=index, column=3).value = meeting._date
                ws.cell(row=index, column=5).value = meeting.jira_ticket
            except Exception as e:  
                print("Error: ",e)
            index+=1
        init_new_rows=total_rows+1
        with_format_hours=(len(mettings)-1)+init_new_rows
        ws.merge_cells(f'D{init_new_rows}:D{with_format_hours}')
        increment=0
        for increment in range(0, len(mettings)):
            index_row=init_new_rows+increment
            for cell in ws[index_row]:
                cell.alignment = alignment
                cell.border = borde
        index_hour=5
        index_hour=round(len(mettings)/2+init_new_rows,0)
        index_hour=int(index_hour)
        ws[f'D{init_new_rows}'] = TOTAL_HOURS    
        wb.save(EXCEL_PATH_READ)
        
    except Exception as e:  
        print("Error: ",e)
        print("Creating File...")       
        writer,worksheet,header_bold_format,content_format=create_file()    
        cell_width=300
        row_index=4
        for i in range(len(items)):
            worksheet.set_column_pixels(i, 0, cell_width)
            worksheet.write(row_index,i,items[i], header_bold_format)
        mettings=get_mettings()
        index=5
        for meeting in mettings:
            worksheet.write(index,0,meeting.title, content_format)
            worksheet.write(index,1,meeting.proyect,content_format)
            worksheet.write(index,2,meeting._date,content_format)
            worksheet.write(index,4,meeting.jira_ticket,content_format)
            index+=1
        index_hour=5
        with_formar_hours=(len(mettings)-1)+index_hour
        total_hour_col=3
        worksheet.merge_range(index_hour, total_hour_col, with_formar_hours, total_hour_col, TOTAL_HOURS, content_format)
        writer.close()

